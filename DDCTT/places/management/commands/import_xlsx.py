"""
Command import_xlsx for places app
"""
from django.core.management import BaseCommand
from openpyxl import load_workbook

from places.models import Place


class Command(BaseCommand):
    """
    class Command for import data from xlsx files to DB
    """

    def handle(self, *args, **options):
        """
        Get info from filename, create objects of place models
        """
        filename = 'places.xlsx'
        wb = load_workbook(filename, read_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, point, raiting = row
            longitude, latitude = point.split(', ')
            point = 'POINT({} {})'.format(longitude, latitude)
            Place.objects.get_or_create(name=name, point=point, raiting=raiting)


